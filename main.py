import openpyxl
import os
from math import sqrt
from prettytable import PrettyTable
import matplotlib.pyplot as plt


def distribution_width_check(list, max, min):
    max_value = None
    min_value = None
    for value in list:
        if value > max+0.1:
            continue
        if value < min-0.1:
            continue
        if not min_value:
            min_value = value
        elif value < min_value:
            min_value = value
        if not max_value:
            max_value = value
        elif value > max_value:
            max_value = value
    return max_value, min_value


def strike(number):
    result = ''
    striked = int((number-(number%5))/5)
    unstriked = number%5
    unstriked_text = 'I'
    text = 'IIII'
    for c in text:
        result = result + c + '\u0336'
    result += ' '
    ret = str(result*striked) + str(unstriked_text*unstriked)
    return ret


def average(list):
    summ = 0
    for i in list:
        summ += i
    return round(summ/len(list),4)


def sum_sq_dev(list, center):
    summ = 0
    for i in list:
        summ = summ+((i-center)**2)
    return round(summ,5)


os.chdir('/home/pixelastra/LR3/')

wb = openpyxl.load_workbook('2.xlsx')
sheet = wb['Лист 1']
parsed_list = []
table = PrettyTable()

limit_min = sheet.cell(row=54, column=2).value
limit_max = sheet.cell(row=55, column=2).value
for row_i in range(3,53):
    parsed_list.append(sheet.cell(row=row_i, column=2).value)
dis_vars = distribution_width_check(parsed_list, limit_max, limit_min)
distribution_width = round(dis_vars[0] - dis_vars[1], 4)
k = round(sqrt(len(parsed_list)))
h = round(distribution_width/k, 2)
av = average(parsed_list)
sum_sq = sum_sq_dev(parsed_list, av)
lim_list = [limit_min]
for i in range(k+2):
    lim_list.append(round(lim_list[i]+h,3))
print(lim_list)

print('\n\n\n\n')
print('Широта распределения R =', distribution_width)
print("Предварительное число интервалов k =", k)
print("Широта интервала h =", h)
print('Рабочий настроечный размер Aр =', (dis_vars[0]+dis_vars[1])*0.5)
print('\n\n\n\n')

table.field_names = ['№ п/п', 'Интервал', 'Значение середины интервала', 'Подсчет частот', 'Частота', 'Накопленная частота']
sum_frequency = 0
freq_sum = []
names_list = []
for i in range(len(lim_list)-1):
    current_lim = lim_list[i]
    next_lim = lim_list[i+1]
    if current_lim <= limit_max:
        text_lim = str(current_lim)+'-'+str(next_lim)
        text_center = round((lim_list[i]+lim_list[i+1])/2,3)
        names_list.append(text_lim)
        frequency = []
        for n in parsed_list:
            if current_lim <= n < next_lim:
                frequency.append(n)
    else:
        continue
    table.add_row([i+1, text_lim, text_center, strike(len(frequency)), len(frequency), sum_frequency+len(frequency)])
    freq_sum.append(len(frequency))
    sum_frequency += len(frequency)
print(table)


print('\n\n\n\n')
print('Среднее арифметическое x =', av)
print('Центр группирования значений M(x) =', average(parsed_list))
print('Сумма квадратов отклонений S=', sum_sq)
print('Дисперсия σx^2=', round(sum_sq/(len(parsed_list)-1), 5))
print('Квадратичное отклонение σx=', round(sqrt(sum_sq/(len(parsed_list)-1)), 5))
print('\n\n\n\n\n')

print(names_list)
print(freq_sum)

names = names_list
values = freq_sum
plt.rc('xtick',labelsize=8)
plt.title('Гистограмма распределения величин')
plt.xlabel('Размеры')
plt.ylabel('Частота')
plt.xticks([0,0.999,1.999,2.994,3.998,5,6,7,7.993,8.992], rotation=14)
plt.bar(names, values, align='edge', width=0.999, edgecolor='black', color='white')
plt.savefig('plot.png')

